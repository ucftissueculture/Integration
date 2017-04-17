# MediaPrep Studio
# UCF Senior Design Project - Blue Agave
# Application written by Clarisse Vamos for Agri-Starts

#Import libraries needed to create application
from tkinter import *
from tkinter import ttk
from time import time as tm
import time

#Import libraries needed for saving to excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import calendar
from pathlib import Path

# Create Application Root
root = Tk()
root.title("MediaPrep Studio")

# Global Variables
mediaType = "None"          #type of media used in process
trayType = "None"           #type of trays run (jars, tubs, test tubes)
processType = "None"        #type of process carried out
startVol = 0                #volume entered by user as starting volume in tank
numTrays = 0                #number of completed trays
fillVol = 0                 #volume of media that has been used to fill jars
t0=0                        #process start time
rtime = 0                   #process run time
finalt = 0                  #final run time, formatted nicely
errormsg = ""               #error message if needed

# The Application Frame
class Application(Frame):

    # Create Initialization Sequence
    def __init__(self, master):
        Frame.__init__(self, master)
        self.create_title()
        self.create_setup()
        self.create_results()
        self.create_buttons()

    # Create Title Frame
    def create_title(self):
        self.tframe = ttk.Frame(root)
        self.title = ttk.Label(self.tframe, text='Agri-Starts Media Preparation', font = "calibri 16 bold")
        self.title.pack()
        self.tframe.config(padding=(10, 10))
        self.tframe.pack()

    # Create Setup Frame
    def create_setup(self):
        self.sframe = ttk.LabelFrame(root, height=200, width=500, text='Setup')
        self.sframe.config(padding=(10, 10))

        #Create the 'Media Type' label
        self.lbl_media = Label(self.sframe, justify=RIGHT, padx=10, text="Media Type:", font="Verdana 10")
        self.lbl_media.grid(column=1,row=1)

        #Media drop down box
        self.media = StringVar(self.sframe)
        self.media.set("None")  # default value
        self.medialist = ttk.Combobox(self.sframe, textvariable=self.media)

        #Change this list to add/subtract media types
        self.medialist['values'] = ('1', '2', '5', '7','10','15','20','30','K1','K3','K5','K7','K10','K1B','K3B','K5B','B1','B5','B10','L','A4','A8','MS','MS 1/2')
        self.medialist.grid(column=2, row=1)

        #Create the 'Media Volume' label
        self.lbl_startVol = Label(self.sframe, justify=RIGHT, padx=10, text="Media Volume (L):", font="Verdana 10")
        self.lbl_startVol.grid(column=1,row=2)

        #Setting the media volume
        self.mstartVol = Entry(self.sframe)
        self.mstartVol.grid(column=2,row=2)

        # Create the 'Tray Type' label
        self.lbl_tray = Label(self.sframe, justify=RIGHT, padx=10, text="Tray Type:", font="Verdana 10")
        self.lbl_tray.grid(column=1, row=3)

        # Process drop down box
        self.trayType = StringVar(self.sframe)
        self.trayType.set("None")  # default value
        self.trayTypelist = OptionMenu(self.sframe, self.trayType, "Jars", "Tubs", "Test Tubes")
        self.trayTypelist.grid(column=2, row=3)

        #Create the 'Process Type' label
        self.lbl_process = Label(self.sframe, justify=RIGHT, padx=10, text="Process Type:", font="Verdana 10")
        self.lbl_process.grid(column=1,row=4)

        #Process drop down box
        self.process = StringVar(self.sframe)
        self.process.set("None")  # default value
        self.processlist = OptionMenu(self.sframe, self.process, "Automated", "Manual", "Testing")
        self.processlist.grid(column=2, row=4)

        #Pack Frame
        self.sframe.pack(fill="both", expand="yes", anchor='w')

    # Create Button Frame
    def create_buttons(self):
        self.bframe = ttk.Frame(root)
        self.bframe.config(height=200, width=500)
        self.bframe.config(padding=(10, 10))

        #Start Button
        self.select = ttk.Button(self.bframe, text='Start Process', command=self.selectClick)
        self.select.config(state=NORMAL)
        self.select.grid(column=1,row=1)

        #Stop Button
        self.stop = ttk.Button(self.bframe, text = 'STOP', command=self.stopClick)
        self.stop.config(state = DISABLED)
        self.stop.grid(column=2,row=1)

        #Pack Frame
        self.bframe.pack()

    #This method handles the clicking of the 'Start Process' button
    def selectClick(self):
        # set current time as t0
        global t0
        t0 = tm()

        # retrieve global version variables that will be used
        global mediaType, startVol, processType,trayType, numTrays, fillVol
        mediaType = self.media.get()
        startVol = self.mstartVol.get()
        processType = self.process.get()
        trayType = self.trayType.get()
        fillVol = numTrays * 30 * 30 / 1000

        # this will help check that a media volume is entered
        startL = len(self.mstartVol.get())

        # check for missing information before processing start button click
        if processType == "None" or ((processType == "Automated" and ((startL==0 or startVol==0) or (mediaType== "None" or trayType=="None"))) or (processType == "Manual" and (mediaType == "None" or trayType == "None"))):
            # create top level pop up window to warn user of missing info
            top = Toplevel()
            top.title("Process Start Failed")
            about_message = "Cannot start process due to missing user selection(s). Please correct before attempting to start media preparation."

            # create the window's message
            msg = Message(top, text=about_message, justify=CENTER, pady=5)
            msg.pack()

            # add a button that allows user to exit window
            button = Button(top, text="Dismiss", command=top.destroy, padx=7, pady=5)
            button.pack()

        else:
            # Change button states
            self.select.config(state=DISABLED)
            self.stop.config(state=NORMAL)

            # Start process
            self.runprocess()

	#This method will run the individual logic of every process
    def runprocess(self):
		#Automated process, so get the starting volume and use that as limit for number of trays that can be run.
        if (processType == "Automated"):
            global startVol
            intVol = int(startVol.replace(',', ''))
            traylimit = intVol * 1000 / 30 / 30
            n = 0
            while (traylimit > 0):
                numTrays = n
                fillVol = float(numTrays) * 30 * 30 / 1000

                traylimit = traylimit - 1
                print("loz" + numTrays.__str__())
                print traylimit.__str__()
                n = n + 1
        elif (processType == "Manual"):
            print ("MADE IT")

        elif (processType == "Testing"):

            numTrays = 300
            fillVol = numTrays * 30 * 30

        else:
            print("Error.")
        trays = 200
        return trays

    def stopClick(self):

        # Get the global version of variables and calculate the run time
        global t0, rtime, finalt, errormsg, processType
        rtime = tm() - t0
        finalt = self.formtime()

		#Automated, so just call the save and display method
        if processType == "Automated":
            save = self.save_and_display()
        
		#Manual, so take in input on number of trays completed and use that to calculate the filled volume. Then call save and diaplay.
		elif processType == "Manual":
            inputwindow = Toplevel(root)
            inputwindow.title("Input Manual Results")
            inputwindow.geometry("%dx%d%+d%+d" % (300, 200, 250, 125))

            lbl_msg = Label(inputwindow, justify=RIGHT, padx=30, pady=20, text="Input type & number of trays completed:", font="Verdana 10")
            lbl_msg.pack()

            # Create the 'Trays Completed' label
            lbl_nTrays = Label(inputwindow, justify=RIGHT, padx=15, text="Trays Completed:", font="Verdana 10")
            lbl_nTrays.pack()

            # Set the number of trays completed
            trayValue = Entry(inputwindow)
            trayValue.pack()

            # Purely aesthetic label
            lbl_pretty = Label(inputwindow, justify=RIGHT, padx=15, text="", font="Verdana 10")
            lbl_pretty.pack()

            resetbtn = Button(inputwindow, text="Done", command=self.save_and_display, padx=10, pady=5)
            resetbtn.pack()

		#Testing, so just reset values after process is stopped.
        elif processType == "Testing":
            self.reset()

		#No process entered, so will throw error
        else:
            errormsg = "No process type."

        # Change button states
        self.stop.config(state=DISABLED)
        self.select.config(state=NORMAL)

	#This method calls the save method and displays the process results to the user
    def save_and_display(self):

        global mediaType, processType, numTrays, fillVol, finalt, errormsg, trayType
        # save the results
        errormsg = self.savedata()

        # Display toplevel window with results
        finalout = Toplevel(root)
        finalout.deiconify()
        finalout.title("Final Process Results")
        finalout.geometry("%dx%d%+d%+d" % (300, 200, 250, 125))

        # Add a title and message to the window
        about_title = "FINAL PROCESS RESULTS\n\n"
        about_message = "Media Type:\t\t" + mediaType + "\nTray Type:\t\t" + trayType + "\nProcess Type:\t\t" + processType + "\nTrays Completed:\t\t" + numTrays.__str__() + "\nFilled Volume (L):\t\t" + fillVol.__str__() + "\nRun Time (H:M:S):\t" + finalt + "\n\n" + errormsg
        msg = Message(finalout, text=about_title + about_message, width=350, anchor='w', pady=10)
        msg.pack()

        # Close the results pop up and reset all values on screen
        resetbtn = Button(finalout, text="Reset Values", command=self.reset, padx=7, pady=5)
        resetbtn.pack()

	#This method formats the run time to H:M:S
    def formtime(self):
        global rtime
        m, s = divmod(rtime, 60)
        h, m = divmod(m, 60)
        ftime = "%02d:%02d:%02d" % (h, m, s)
        return ftime

	#This method will save the results data to an Excel sheet
    def savedata(self):
        print "SAVING...."
        # Get current month and year info
        monthnum = datetime.now().month
        year = datetime.now().year
        month = calendar.month_name[monthnum]

        # get global version of variables to save
        global mediaType, processType, numTrays, fillVol, rtime, trayType

        # Create name of data file, corresponding to current year and month
        filename = "MediaPrepData_" + month.__str__() + year.__str__() + ".xlsx"
        filepathend = "\MediaPrepData_" + month.__str__() + year.__str__() + ".xlsx"
        filepath = Path("C:\Users\Clarisse\Documents\MediaPrepStudio" + filepathend)

        try:
            # Open file if it exists in path, otherwise create
            if filepath.is_file():
                # Open workbook for the month
                currdatalog = openpyxl.load_workbook(filename)
                currsheet = currdatalog.get_sheet_by_name("Media Prep Log")

                # Find next available row in file
                rownum = currsheet.max_row + 1

                # Print data to that location
                currsheet.cell(row=rownum, column=1).value = "Date..."
                currsheet.cell(row=rownum, column=2).value = "Time..."
                currsheet.cell(row=rownum, column=3).value = mediaType
                currsheet.cell(row=rownum, column=4).value = trayType
                currsheet.cell(row=rownum, column=5).value = processType
                currsheet.cell(row=rownum, column=6).value = numTrays
                currsheet.cell(row=rownum, column=7).value = fillVol
                currsheet.cell(row=rownum, column=9).value = self.formtime()

                # Save the file with updated data
                currdatalog.save(filename)
                print "SAVING COMPLETE."

            # File not found, so create one
            else:
                print "File not found...creating one."

                # Create workbook for the month
                datalog = Workbook()

                # Create worksheet and give it the proper name
                sheet = datalog.active
                sheet.title = "Media Prep Log"
                print datalog.sheetnames

                # Create heading for file
                sheet.cell(row=1, column=1).value = "Media Prep Log"
                sheet.cell(row=1, column=1).value = "Agri-Starts Media Preparation Data Log"
                sheet.cell(row=2, column=1).value = "Month:"
                sheet.cell(row=2, column=2).value = month.__str__()
                sheet.cell(row=3, column=1).value = "Year:"
                sheet.cell(row=3, column=2).value = year
                sheet.cell(row=3, column=2).alignment = Alignment(horizontal='left')

                # Now create table headers for data
                sheet.cell(row=5, column=1).value = "Date"
                sheet.cell(row=5, column=2).value = "Time"
                sheet.cell(row=5, column=3).value = "Media Type"
                sheet.cell(row=5, column=4).value = "Tray Type"
                sheet.cell(row=5, column=5).value = "Process Type"
                sheet.cell(row=5, column=6).value = "Trays Completed"
                sheet.cell(row=5, column=7).value = "Filled Volume (L)"
                sheet.cell(row=5, column=8).value = "Run Time (H:M:S)"

                # Set column widths
                sheet.column_dimensions['A'].width = 12.0
                sheet.column_dimensions['B'].width = 12.0
                sheet.column_dimensions['C'].width = 13.0
                sheet.column_dimensions['D'].width = 13.0
                sheet.column_dimensions['E'].width = 13.0
                sheet.column_dimensions['F'].width = 16.0
                sheet.column_dimensions['G'].width = 15.0
                sheet.column_dimensions['H'].width = 16.0

                # Set row number to 6, which is start of data
                rownum=6

                # Print data to that location
                sheet.cell(row=rownum, column=1).value = "Date..."
                sheet.cell(row=rownum, column=2).value = "Time..."
                sheet.cell(row=rownum, column=3).value = mediaType
                sheet.cell(row=rownum, column=4).value = trayType
                sheet.cell(row=rownum, column=5).value = processType
                sheet.cell(row=rownum, column=6).value = numTrays
                sheet.cell(row=rownum, column=7).value = fillVol
                sheet.cell(row=rownum, column=8).value = self.formtime()

                # Save file
                datalog.save(filename)
                print "SAVING COMPLETE."
		
		#If there is a problem saving the file, an error will be returned and shown to the user
        except Exception:
            error = "Problem saving to file. Please manually enter data into file if you would like it to be logged"
            return error
        
		#Will return empty message if no error
		return ""

	#This performs a reset of all values on the main application window
    def reset(self):
        global trayType
        trayType = "None"

        # Reset Input Values
        self.media.set("None")
        mediaType = "None"
		trayType = "None"
        self.mstartVol.delete(0, 'end')
        self.process.set("None")
        processType = "None"

        # Reset timer
        global rtime
        rtime = 0


# End
app = Application(root)
root.mainloop()
