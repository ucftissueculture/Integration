# MediaPrep Studio
# UCF Senior Design Project - Blue Agave
# Application written by Clarisse Vamos for Agri-Starts

#Import libraries needed to create application
from tkinter import *
from tkinter import ttk
from time import time as tm
import time

#Import libraries needed for saving to excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import calendar
from pathlib import Path


# Create Application Root
root = Tk()
root.title("MediaPrep Studio")

# Change Style of Certain Elements
s = ttk.Style()
s.configure('TLabelframe.Label', font='calibri 13 bold')
s.configure('TButton', font='calibri 14 bold')

# Global Variables
mediaType = "None"          #type of media used in process
processType = "None"        #type of process carried out
startVol = 0                #volume entered by user as starting volume in tank
numTrays = 0                #number of completed trays
fillVol = 0                 #volume of media that has been used to fill jars
t0=0                        #process start time
rtime = 0                   #process run time

#Create and hide toplevel window for results
finalout = Toplevel()
finalout.withdraw()

# The Application Frame
class Application(Frame):

    # Create Initialization Sequence
    def __init__(self, master):
        Frame.__init__(self, master)
        self.create_title()
        self.create_setup()
        self.create_results()
        self.create_buttons()

    # Create Title Frame
    def create_title(self):
        self.tframe = ttk.Frame(root)
        self.title = ttk.Label(self.tframe, text='Agri-Starts Media Preparation', font = "calibri 16 bold")
        self.title.pack()
        self.tframe.config(padding=(10, 10))
        self.tframe.pack()

    # Create Setup Frame
    def create_setup(self):
        self.sframe = ttk.LabelFrame(root, height=200, width=500, text='Setup')
        self.sframe.config(padding=(10, 10))

        #Create the 'Media Type' label
        self.lbl_media = Label(self.sframe, justify=RIGHT, padx=10, text="Media Type:", font="Verdana 10")
        self.lbl_media.grid(column=1,row=1)

        #Media drop down box
        self.media = StringVar(self.sframe)
        self.media.set("None")  # default value
        self.medialist = ttk.Combobox(self.sframe, textvariable=self.media)

        #Change this list to add/subtract media types
        self.medialist['values'] = ('1', '2', '5', '7','10','15','20','30','K1','K3','K5','K7','K10','K1B','K3B','K5B','B1','B5','B10','L','A4','A8','MS','MS 1/2')
        self.medialist.grid(column=2, row=1)

        #Create the 'Media Volume' label
        self.lbl_startVol = Label(self.sframe, justify=RIGHT, padx=10, text="Media Volume (L):", font="Verdana 10")
        self.lbl_startVol.grid(column=1,row=2)

        #Setting the media volume
        self.mstartVol = Entry(self.sframe)
        self.mstartVol.grid(column=2,row=2)

        #Create the 'Process Type' label
        self.lbl_process = Label(self.sframe, justify=RIGHT, padx=10, text="Process Type:", font="Verdana 10")
        self.lbl_process.grid(column=1,row=3)

        #Process drop down box
        self.process = StringVar(self.sframe)
        self.process.set("None")  # default value
        self.processlist = OptionMenu(self.sframe, self.process, "Automated", "Manual", "Testing")
        self.processlist.grid(column=2, row=3)

        #Pack Frame
        self.sframe.pack(fill="both", expand="yes", anchor='w')

    # Create Current Process Frame
    def create_results(self):
        self.rframe = ttk.LabelFrame(root, height=200, width=500, text='Current Process')
        self.rframe.config(padding=(10, 10))

        # Create the 'Media Type' label
        self.lbl_media2 = Label(self.rframe, justify=RIGHT, padx=10, text="Media Type:", font="Verdana 10")
        self.lbl_media2.grid(column=1,row=1)

        # Create Media Type value label
        self.mediaValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.mediaValue.grid(column=2,row=1)

        # Create the 'Process Type' label
        self.lbl_process2 = Label(self.rframe, justify=RIGHT, padx=10, text="Process Type:", font="Verdana 10")
        self.lbl_process2.grid(column=1, row=2)

        # Create Process Type value label
        self.processValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.processValue.grid(column=2, row=2)

        # Create the 'Trays Completed' label
        self.lbl_numTrays = Label(self.rframe, justify=RIGHT, padx=10, text="Trays Completed:", font="Verdana 10")
        self.lbl_numTrays.grid(column=1,row=3)

        # Create Trays Completed value label
        self.trayValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.trayValue.grid(column=2, row=3)

        # Create the 'Filled Volume' label
        self.lbl_fillVol = Label(self.rframe, justify=RIGHT, padx=10, text="Filled Volume (L):", font="Verdana 10")
        self.lbl_fillVol.grid(column=1,row=4)

        # Create Filled Volume value label
        self.fillValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.fillValue.grid(column=2, row=4)

        #Pack Frame
        self.rframe.pack(fill="both", expand="yes",anchor='w')

    # Create Button Frame
    def create_buttons(self):
        self.bframe = ttk.Frame(root)
        self.bframe.config(height=200, width=500)
        self.bframe.config(padding=(10, 10))

        #Start Button
        self.select = ttk.Button(self.bframe, text='Start Process', command=self.selectClick)
        self.select.config(state=NORMAL)
        self.select.grid(column=1,row=1)

        #Stop Button
        self.stop = ttk.Button(self.bframe, text = 'STOP', command=self.stopClick)
        self.stop.config(state = DISABLED)
        self.stop.grid(column=2,row=1)

        #Pack Frame
        self.bframe.pack()

    # Handle the clicking of the 'Start Process' button
    def selectClick(self):
        # set current time as t0
        global t0
        t0 = tm()

        # retrieve global version variables that will be used
        global mediaType, startVol, processType, numTrays, fillVol
        mediaType = self.media.get()
        startVol = self.mstartVol.get()
        processType = self.process.get()
        fillVol = numTrays * 30 * 30 / 1000

        # this will help check that a media volume is entered
        startL = len(self.mstartVol.get())

        # check for missing information before processing start button click
        if mediaType== "None" or startL==0 or startVol==0 or processType== "None":
            # create top level pop up window to warn user of missing info
            top = Toplevel()
            top.title("Process Start Failed")
            about_message = "Cannot start process due to missing user selection(s). Please correct before attempting to start media preparation."

            # create the window's message
            msg = Message(top, text=about_message, justify=CENTER, pady=10)
            msg.pack()

            # add a button that allows user to exit window
            button = Button(top, text="Dismiss", command=top.destroy, padx=7, pady=5)
            button.pack()
        else:
            # print values to "Current Process"
            self.mediaValue["text"] = mediaType
            self.processValue["text"] = processType

            # Change button states
            self.select.config(state=DISABLED)
            self.stop.config(state=NORMAL)

            # Start process
            self.runprocess()

    def reset(self):

        #destroy the results pop up
        finalout.destroy()

        # Reset Input Values
        self.media.set("None")
        mediaType = "None"
        self.mstartVol.delete(0, 'end')
        self.process.set("None")
        processType = "None"

        # Reset Output Values
        self.mediaValue["text"] = ""
        self.processValue["text"] = ""
        self.trayValue["text"] = ""
        self.fillValue["text"] = ""
        self.rtValue = "00:00:00"

        # Reset timer
        global rtime
        rtime = 0

    def formtime(self):
        global rtime
        m, s = divmod(rtime, 60)
        h, m = divmod(m, 60)
        ftime = "%02d:%02d:%02d" % (h, m, s)
        return ftime

    def savedata(self):
        print "SAVING...."
        # Get current month and year info
        monthnum = datetime.now().month
        year = datetime.now().year
        month = calendar.month_name[monthnum]

        # Create name of data file, corresponding to current year and month
        filename = "MediaPrepData_" + month.__str__() + year.__str__() + ".xlsx"
        filepathend = "\MediaPrepData_" + month.__str__() + year.__str__() + ".xlsx"
        filepath = Path("C:\Users\Clarisse\Documents\MediaPrepStudio" + filepathend)

        try:
            # Open file if it exists in path, otherwise create
            if filepath.is_file():
                # Open workbook for the month
                currdatalog = openpyxl.load_workbook(filename)
                currsheet = currdatalog.get_sheet_by_name("Media Prep Log")

                # Find next available row in file
                rownum = currsheet.max_row + 1

                #get global version of variables to save
                global mediaType, processType,numTrays,fillVol, rtime

                # Print data to that location
                currsheet.cell(row=rownum, column=1).value = "Date..."
                currsheet.cell(row=rownum, column=2).value = "Time..."
                currsheet.cell(row=rownum, column=3).value = mediaType
                currsheet.cell(row=rownum, column=4).value = processType
                currsheet.cell(row=rownum, column=5).value = numTrays
                currsheet.cell(row=rownum, column=6).value = fillVol
                currsheet.cell(row=rownum, column=7).value = self.formtime()

                # Save the file with updated data
                currdatalog.save(filename)
                print "SAVING COMPLETE."

            # File not found, so create one
            else:
                print "File not found...creating one."

                # Create workbook for the month
                datalog = Workbook()

                # Create worksheet and give it the proper name
                sheet = datalog.active
                sheet.title = "Media Prep Log"
                print datalog.sheetnames

                # Create heading for file
                sheet.cell(row=1, column=1).value = "Media Prep Log"
                sheet.cell(row=1, column=1).value = "Agri-Starts Media Preparation Data Log"
                sheet.cell(row=2, column=1).value = "Month:"
                sheet.cell(row=2, column=2).value = month.__str__()
                sheet.cell(row=3, column=1).value = "Year:"
                sheet.cell(row=3, column=2).value = year
                sheet.cell(row=3, column=2).alignment = Alignment(horizontal='left')

                # Now create table headers for data
                sheet.cell(row=5, column=1).value = "Date"
                sheet.cell(row=5, column=2).value = "Time"
                sheet.cell(row=5, column=3).value = "Media Type"
                sheet.cell(row=5, column=4).value = "Process Type"
                sheet.cell(row=5, column=5).value = "Trays Completed"
                sheet.cell(row=5, column=6).value = "Filled Volume (L)"
                sheet.cell(row=5, column=7).value = "Run Time (H:M:S)"

                # Set column widths
                sheet.column_dimensions['A'].width = 12.0
                sheet.column_dimensions['B'].width = 12.0
                sheet.column_dimensions['C'].width = 13.0
                sheet.column_dimensions['D'].width = 13.0
                sheet.column_dimensions['E'].width = 16.0
                sheet.column_dimensions['F'].width = 15.0
                sheet.column_dimensions['G'].width = 16.0

                # Set row number to 6, which is start of data
                rownum=6

                # Print data to that location
                sheet.cell(row=rownum, column=1).value = "Date..."
                sheet.cell(row=rownum, column=2).value = "Time..."
                sheet.cell(row=rownum, column=3).value = mediaType
                sheet.cell(row=rownum, column=4).value = processType
                sheet.cell(row=rownum, column=5).value = numTrays
                sheet.cell(row=rownum, column=6).value = fillVol
                sheet.cell(row=rownum, column=7).value = self.formtime()

                # Save file
                datalog.save(filename)
                print "SAVING COMPLETE."

        except Exception:
            error = "Problem saving to file. Please manually enter data into file if you would like it to be logged"
            return error
        return ""

    def stopClick(self):

        #calculate run time
        global t0, rtime
        rtime = tm() - t0
        finalt = self.formtime()

        #get final output from global variables already calculated
        global mediaType
        global processType
        global numTrays
        global fillVol
        global sec

        # save the results
        errormsg = self.savedata()

        # Display toplevel window with results
        global finalout
        finalout.deiconify()
        finalout.title("Final Process Results")
        finalout.geometry("%dx%d%+d%+d" % (300, 200, 250, 125))

        # Add a title and message to the window
        about_title = "FINAL PROCESS RESULTS\n\n"
        about_message = "Media Type:\t\t" + mediaType + "\nProcess Type:\t\t" + processType + "\nTrays Completed:\t\t" + numTrays.__str__() + "\nFilled Volume (L):\t\t" + fillVol.__str__() + "\nRun Time (H:M:S):\t" + finalt + "\n\n" + errormsg
        msg = Message(finalout, text=about_title + about_message, width=350, anchor='w', pady=10)
        msg.pack()

        # Close the results pop up and reset all values on screen
        resetbtn = Button(finalout, text="Exit and Reset Values", command= self.reset, padx=7, pady=5)
        resetbtn.pack()

        # Change button states
        self.stop.config(state=DISABLED)
        self.select.config(state=NORMAL)

    # change to actual process method
    def runprocess(self):

        if (processType == "Automated"):
            self.trayValue["text"] = "Lolz"
            global startVol
            intVol = int(startVol.replace(',',''))
            traylimit = intVol * 1000 / 30 / 30
            n = 0
            while (traylimit > 0):
                numTrays = n
                fillVol = float(numTrays) * 30 * 30 / 1000

                self.trayValue["text"] = numTrays.__str__()
                self.fillValue["text"] = fillVol.__str__()

                traylimit = traylimit - 1
                print("loz" + numTrays.__str__())
                print traylimit.__str__()
                n = n + 1

        elif (processType == "Manual"):
            self.trayValue["text"] = "Lolz 2"
            num = 10

            for counter in range(num):
                numTrays = num
                fillVol = numTrays * 30 * 30 / 1000
                self.trayValue["text"] = numTrays.__str__()
                self.fillValue["text"] = fillVol.__str__()

            print ("MADE IT")

        elif (processType == "Testing"):
            self.trayValue["text"] = "Lolz 3"

            numTrays = 300
            fillVol = numTrays * 30 * 30

        else:
            print("Error.")
        trays = 200
        return trays

# End
app = Application(root)
root.mainloop()
